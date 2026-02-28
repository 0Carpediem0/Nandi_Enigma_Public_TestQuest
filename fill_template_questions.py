# -*- coding: utf-8 -*-
"""
Заполняет в kb_test.xlsx колонку tags шаблонными вопросами для каждой темы.
Шаблонные вопросы — примеры запросов пользователя, которые могут привести к этой теме.
"""
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Установите openpyxl: pip install openpyxl")
    raise SystemExit(1)

KB_XLSX = Path(__file__).resolve().parent / "kb_test.xlsx"
if not KB_XLSX.exists():
    KB_XLSX = Path(__file__).resolve().parent.parent / "kb_test.xlsx"


def _normalize_header(h):
    if h is None:
        return ""
    return str(h).strip().lower().replace(" ", "_")


def generate_template_questions(title: str, category: str = "") -> list[str]:
    """По заголовку темы формирует список шаблонных вопросов, которые могут к ней привести."""
    t = (title or "").strip()
    if not t:
        return []
    questions = []
    t_lower = t.lower()

    # Сам заголовок часто уже формулировка темы
    questions.append(t)

    # По ключевым словам добавляем типичные формулировки пользователей
    if "назначение" in t_lower or "назначен" in t_lower:
        questions.extend([
            "Для чего это нужно?",
            "Что это такое?",
        ])
    if "органы управления" in t_lower or "управлен" in t_lower:
        questions.extend([
            "Где кнопки управления?",
            "Как управлять?",
        ])
    if "частота" in t_lower and "миган" in t_lower:
        questions.extend([
            "Как часто мигает?",
            "Частота мигания индикатора",
        ])
    if "звуковая сигнализация" in t_lower or "нештатн" in t_lower:
        questions.extend([
            "Что делать при срабатывании звуковой сигнализации?",
            "Как отключить звуковой сигнал?",
        ])
    if "настройка" in t_lower and ("адрес" in t_lower or "скорост" in t_lower or "rs-232" in t_lower or "rs-485" in t_lower):
        questions.extend([
            "Как настроить сетевой адрес?",
            "Как изменить скорость RS-485?",
        ])
    if "уровн" in t_lower and "доступ" in t_lower or "пароль" in t_lower:
        questions.extend([
            "Как установить пароль?",
            "Где сменить пароль доступа?",
        ])
    if "режим обслуживания" in t_lower or "обслуживание канала" in t_lower:
        questions.extend([
            "Как включить режим обслуживания?",
            "Режим обслуживания канала",
        ])
    if "неисправность" in t_lower or "авария" in t_lower or "светит" in t_lower:
        questions.extend([
            "Что делать если светит авария?",
            "Почему горит авария?",
            "Как снять аварию?",
        ])
    if "сброс" in t_lower and "программ" in t_lower:
        questions.extend([
            "Как сбросить программу?",
            "Как сделать сброс настроек?",
        ])
    if "периодичность" in t_lower or "техническое обслуживание" in t_lower:
        questions.extend([
            "Как часто проводить обслуживание?",
            "Когда делать ТО?",
        ])
    if "очистка" in t_lower and "архив" in t_lower:
        questions.extend([
            "Как очистить архив?",
            "Как удалить старые данные?",
        ])
    if "кабель" in t_lower or "подключен" in t_lower and "датчик" in t_lower:
        questions.extend([
            "Какой кабель нужен для датчика?",
            "Требования к кабелю подключения",
        ])

    # Универсальные варианты, если мало набралось
    if len(questions) <= 1:
        if "как" in t_lower:
            questions.append("Как это сделать?")
        else:
            questions.append(f"Как {t_lower}?")
        questions.append(f"Расскажите про {t_lower}")

    # Убираем дубликаты с сохранением порядка
    seen = set()
    unique = []
    for q in questions:
        q_clean = q.strip()
        if q_clean and q_clean not in seen:
            seen.add(q_clean)
            unique.append(q_clean)
    return unique[:8]


def main():
    if not KB_XLSX.exists():
        print(f"Файл не найден: {KB_XLSX}")
        return 1

    wb = openpyxl.load_workbook(KB_XLSX)
    ws = wb.active
    if not ws:
        print("Нет активного листа.")
        return 1

    headers = [_normalize_header(cell.value) for cell in ws[1]]
    title_idx = -1
    category_idx = -1
    tags_idx = -1
    for i, h in enumerate(headers):
        if h in ("title", "заголовок", "вопрос", "question", "question_template"):
            title_idx = i
        if h in ("category", "категория"):
            category_idx = i
        if h in ("tags", "теги"):
            tags_idx = i
    if title_idx < 0:
        for i, h in enumerate(headers):
            if "title" in h or "вопрос" in h or "question" in h:
                title_idx = i
                break
    if title_idx < 0:
        print("Не найдена колонка с заголовком темы (title, вопрос).")
        return 1
    if tags_idx < 0:
        print("Не найдена колонка tags (или теги).")
        return 1

    tags_col = tags_idx + 1

    filled = 0
    for row_idx in range(2, ws.max_row + 1):
        title_val = ws.cell(row=row_idx, column=title_idx + 1).value
        title = (title_val if isinstance(title_val, str) else str(title_val or "")).strip()
        if not title:
            continue
        category_val = ""
        if category_idx >= 0:
            cat_cell = ws.cell(row=row_idx, column=category_idx + 1).value
            category_val = (cat_cell if isinstance(cat_cell, str) else str(cat_cell or "")).strip()
        questions = generate_template_questions(title, category_val)
        ws.cell(row=row_idx, column=tags_col, value="; ".join(questions))
        filled += 1

    wb.save(KB_XLSX)
    print(f"Заполнены шаблонные вопросы в колонке tags для {filled} тем в {KB_XLSX}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
